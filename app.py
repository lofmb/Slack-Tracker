"""
The Slack side of the workshop tracker: what an assembler sees, and what happens
when they press it.

This file owns the conversation with the assembler. It builds the working card and
the forms, decides which buttons a job should be offering, reads what comes
back, and says no in words when a job cannot do what was asked. It holds no
data of its own: every durable operation goes to database.py, which is the
other half of the tracker and talks to LMSA.

A handler here reads the same way throughout:

    the assembler presses something
        -> work out which job, and whether it is theirs to touch
            -> check what the job actually allows
                -> ask database.py to record it
                    -> rebuild the card so it shows the new truth

The job runs Field -> Border -> Packing. Field and Border each carry setup and
sheeting underneath them, cutting is measured inside sheeting, and exactly one
item of work accrues at a time. Those rules belong to the job, not to this
file; the sections below say where each part of the conversation lives.
"""

import os
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler
from slack_sdk.errors import SlackApiError
from io import BytesIO

# Importing database functions

import database

# Loading environment variables from .env
load_dotenv()

# Initialsing Slack with the Bot Token and Signing Secret

app = App(
    token=os.environ.get("SLACK_bot_token"),
    signing_secret=os.environ.get("SLACK_SIGNING_SECRET"),
    # Bolt runs the handlers on worker threads. This executor carries the
    # "which Slack delivery is this" note set by the middleware below onto
    # the worker thread, so a redelivered click is still recognised as the
    # same action there. See listener_executor() in database.py.
    listener_executor=database.listener_executor()
)

# ---------------------------------------------------------------------------
# Forms and wording shared by more than one route
# ---------------------------------------------------------------------------
# Defined once here because two different paths reach each of them, and a
# change to one must not miss the other.

def notes_modal_view(metadata, summary_text):
    """
    The last form: what happened, before the job closes.

    Built here rather than inline so it reads beside the packing form it
    follows, and so the two forms that end a job are changed in one place.
    """
    return {
        "type": "modal",
        "callback_id": "trk_notes_modal",
        "title": {"type": "plain_text", "text": "Finish the job"},
        "submit": {"type": "plain_text", "text": "Finish the job"},
        "close": {"type": "plain_text", "text": "Cancel"},
        "private_metadata": metadata,
        "blocks": [
            {
                "type": "section",
                "text": {"type": "mrkdwn", "text": summary_text}
            },
            {
                "type": "input",
                "block_id": "notes_block",
                "optional": True,
                "label": {"type": "plain_text", "text": "Notes"},
                "element": {
                    "type": "plain_text_input",
                    "multiline": True,
                    "action_id": "general_notes",
                    "placeholder": {"type": "plain_text",
                                    "text": "Anything the next person should know"}
                }
            },
            {
                "type": "input",
                "block_id": "issues_block",
                "optional": True,
                "label": {"type": "plain_text", "text": "Anything go wrong?"},
                "element": {
                    "type": "plain_text_input",
                    "multiline": True,
                    "action_id": "issues",
                    "placeholder": {"type": "plain_text",
                                    "text": "Breakages, wrong material, missing pieces"}
                }
            }
        ]
    }


def border_time_display(task):
    """
    What to show wherever a border time would go.

    A job with no border has no border time. Printing "0 h 0m 0s" would read as
    a border that was worked and happened to take no time, which is a different
    thing and the wrong thing. Every border time on a card, in the summary and
    in the export goes through here so they all say it the same way.
    """
    if task.get("border_skipped"):
        return "No Border"
    return database.format_duration(task["border_elapsed"] or 0)


# ---------------------------------------------------------------------------
# The card the assembler works from
# ---------------------------------------------------------------------------
#
# ONE card, built from what the job actually is, so every route through the
# workflow shows the same thing and a change to it cannot land on one path and
# miss another. It answers four questions, in the order an assembler asks them:
#
#     Which job is this?          the header
#     What am I working on?       the status line
#     How long have I recorded?   the times
#     What can I do next?         the buttons
#
# Everything else the job knows is grouped underneath, because giving every
# field the same weight is what turns a card into a wall of text.

# What each item of work is called wherever the assembler sees it. Setup and
# sheeting share a lane, so a name has to say both which lane and which work.
WORK_NAMES = {
    ("field_sheeting", "setup"): "Field setup",
    ("field_sheeting", "production"): "Field sheeting",
    ("border_sheeting", "setup"): "Border setup",
    ("border_sheeting", "production"): "Border sheeting",
    ("packing", "production"): "Packing",
}

# The lane itself, as opposed to an item of work inside it. Used where a figure
# covers the whole lane - its setup and its sheeting together - so that number
# is never labelled with the name of one of its halves.
LANE_NAMES = {
    "field_sheeting": "Field",
    "border_sheeting": "Border",
    "packing": "Packing",
}

# The press that says a lane is genuinely done.
FINISH_LABELS = {
    "field_sheeting": "Field sheeting finished",
    "border_sheeting": "Border finished",
    "packing": "Packing finished",
}

# What state a job is in, as a mark an assembler reads from three feet away without
# reading anything. The card and the workshop channel use the same four, so a
# job that is running looks the same wherever it is seen; changing one of these
# changes both, which is the point of them being here rather than inline.
MARK_RUNNING = "🟢"
# The pause bars are the one of the four that defaults to TEXT presentation, so
# they need the variation selector after them or they draw as a thin monochrome
# glyph beside three coloured ones - or, where the font has no such glyph, as a
# box. The others carry their own emoji presentation already.
MARK_PAUSED = "⏸️"
MARK_CUTTING = "✂️"
MARK_FINISHED = "✅"


def work_name(phase, activity, part=None):
    """
    What to call an item of work, in the words the card uses.

    `part` prefixes the part when the job has more than one, because on a job
    drawn as three parts "Border sheeting" names three different jobs of work
    and the assembler has to be told which. On a single-part job the prefix would
    be noise - there is nothing to tell it apart from - so callers pass None.
    """
    name = WORK_NAMES.get((phase, activity), "Work")
    if part is None:
        return name
    return f"Part {part} {name}"


def lower_name(name):
    """
    A work name as it reads inside a sentence: "start Field sheeting".

    "Part", "Field" and "Border" keep their capitals: they are the names of
    the things the assembler works on, not words in the sentence, and the
    workshop writes them that way. Only a plain activity word - packing - is
    lowercased to sit inside a sentence.
    """
    if name.startswith(("Part ", "Field", "Border")):
        return name
    return name[0].lower() + name[1:]


def part_label(task, part):
    """
    "Part 2", or nothing at all on a job drawn as one part.

    Every naming decision on the card goes through this, so a single-part job
    reads exactly as it did before parts existed and a multi-part one says
    which part throughout.
    """
    if part is None or (task.get("part_count") or 1) < 2:
        return None
    return part


def work_value(task_id, part=None, phase=None, activity=None):
    """
    What a button carries.

    A bare number still means "this job", which is what Edit, Delete and the
    rest need. A button that moves the assembler onto a particular item of work
    names it in full - the job, the part, the lane and the activity - so the
    handler never has to guess which of a job's three borders was meant. The
    two job-level phases, the opening setup and packing, carry an empty part.
    """
    if phase is None:
        return str(task_id)
    return f"{task_id}|{'' if part is None else part}|{phase}|{activity}"


def read_work_value(raw):
    """
    Read a button's value back. Returns (task_id, part, phase, activity).

    Three shapes, because a card posted by an earlier version of the tracker is
    still sitting in somebody's DM and still clickable:
      "12"                            the job, nothing else
      "12|border_sheeting|production"  a lane, before parts were named
      "12|2|border_sheeting|production"   a lane on a part
    The middle one answers None for the part, which every caller then resolves
    to the part the cursor is on - the only part such a card could have meant.
    """
    fields = str(raw).split("|")
    task_id = int(fields[0])
    if len(fields) == 4:
        part = int(fields[1]) if fields[1] else None
        return task_id, part, fields[2], fields[3]
    if len(fields) == 3:
        return task_id, None, fields[1], fields[2]
    return task_id, None, None, None


def lanes_of(task, part):
    """One part's two lanes, as {"field": {...}, "border": {...}}."""
    for row in task.get("parts") or []:
        if row.get("part") == part:
            return row
    return {}


def lane_of(task, part, phase):
    """
    One lane on one part.

    Packing is the job's and is not on a part, so it answers from the flat
    keys - the only lane that does.
    """
    if phase == "packing":
        return {
            "phase": "packing",
            "part": None,
            "present": True,
            "state": task.get("packing_state"),
            "design": None,
            "difficulty": None,
            "setup_elapsed": 0,
            "production_elapsed": task.get("packing_elapsed") or 0,
            "elapsed": task.get("packing_elapsed") or 0,
            "cutting_elapsed": 0,
            "jigs": "",
            "jig_records": [],
        }
    which = "field" if phase == "field_sheeting" else "border"
    return lanes_of(task, part).get(which) or {}


def lane_state(task, part, phase):
    return lane_of(task, part, phase).get("state")


def lane_open(task, part, phase):
    """A lane still to be worked: not finished, and not declared absent."""
    return lane_state(task, part, phase) not in ("complete", "skipped")


def work_elapsed(task, part, phase, activity):
    """
    Seconds recorded against one item of work.

    Packing has no setup, so asking for its setup is nought - not the packing
    total over again. Anything that adds a lane's two activities together
    depends on that.
    """
    lane = lane_of(task, part, phase)
    if phase == "packing":
        return (lane.get("production_elapsed") or 0) if activity == "production" else 0
    return lane.get(f"{activity}_elapsed") or 0


def every_work(task):
    """
    Every item of work this job contains, in the order the assembler reads it.

    Each entry is (part, phase, activity). Packing comes last and belongs to
    no part. This is the one place the job's shape is enumerated, so nothing
    downstream has to know how many parts there are or which lanes they have.
    """
    out = []
    for row in task.get("parts") or []:
        part = row.get("part")
        for phase in ("field_sheeting", "border_sheeting"):
            out.append((part, phase, "setup"))
            out.append((part, phase, "production"))
    out.append((None, "packing", "production"))
    return out


def switch_destinations(task):
    """
    The work the assembler may move to from here.

    Derived from the job rather than listed per card. A lane that is finished,
    or that the assembler declared did not happen, is not somewhere to go; a border
    nobody has described yet has not been reached; and packing has no setup.
    Whatever survives is offered, and the assembler's press is what chooses.
    """
    here = task.get("working_on") or {}
    out = []
    for part, phase, activity in every_work(task):
        # The LANE the assembler is on is not somewhere to move to, whichever of its
        # two activities they are holding. Comparing the activity as well left a
        # lane offering itself: in the first seconds of a sheeting run nothing
        # has accrued yet, so the lane still reads as being at its setup, and
        # the card put up a button that would have paused the very work in hand.
        if phase == here.get("phase") and part == here.get("part"):
            continue
        if not lane_open(task, part, phase):
            continue
        # A lane is described on the way in now, not before it can be reached,
        # so an undescribed border is somewhere to go - the form opens when the
        # assembler gets there. What is NOT somewhere to go is a lane the diagram
        # does not have, which lane_open has already dropped.
        out.append({"part": part, "phase": phase, "activity": activity})
    return out


def other_work_rows(task, already_offered=()):
    """
    The rest of the job, grouped the way the diagram is: one row per part.

    The job is not a wizard. An assembler preparing Part 1's field may need Part 2's
    border first, may come back, may pack in between - and none of that
    finishes what they moved away from. Each row is that part's available
    work, as buttons: the press IS the choice.

    ONE BUTTON PER LANE, not one per activity. Offering a setup and a sheeting
    for every lane of every part put twelve buttons on a three-part card, most
    of them saying the same two words - a wall to read at a bench, and on a
    phone every one of them wrapped. A lane is somewhere to go; the choice
    between preparing it and sheeting it belongs where the assembler already is,
    which is the row above, where "Start field sheeting" already sits next to
    "Pause setup".

    The button drops the part's name, because the row it is in already carries
    it: "Part 2" over "Field" and "Border" reads; "Part 2" over "Part 2 field
    setup" is the same words twice.

    The control says everything about the lane that an assembler needs here. "Field"
    is a lane nobody has opened; "Back to field" is one with time already on it.
    That is the whole of what a line of prose above the row used to say, and it
    is read in the same glance as the press rather than above it.

    NOTHING HERE IS EVER THE HIGHLIGHTED BUTTON. Every press in this section
    pauses the work the assembler is holding, and a card should not colour in the
    one move that stops what they are doing.

    Grouped per part for two reasons, one of them structural. Reading, as
    above. And Slack requires an action_id to be unique WITHIN ITS BLOCK - Part
    1's field and Part 2's field share one, so side by side in a single block
    the card would not render at all. A job drawn as one part has no such
    clash and no headings to sit under, so its rows are merged into one.

    Returns [(heading or None, [buttons])].
    """
    task_id = task["task_id"]
    multi = (task.get("part_count") or 1) > 1

    by_part = {}
    for destination in switch_destinations(task):
        part, phase, activity = destination["part"], destination["phase"], destination["activity"]
        # One entry per LANE. The activity the button carries is where that
        # lane is up to: its setup when nothing has been done on it, and its
        # sheeting once there is sheeting to come back to.
        if activity != lane_entry_activity(task, part, phase):
            continue
        if work_value(task_id, part, phase, activity) in already_offered:
            continue
        recorded = (work_elapsed(task, part, phase, "setup")
                    + work_elapsed(task, part, phase, "production"))
        if phase == "packing":
            # Packing is the JOB's, not a part's, so its button says so rather
            # than leaning on a heading to explain what it belongs to. "Back to
            # pack the job" is not English, so this one names its own return.
            label = "Back to packing" if recorded else "Pack the job"
        else:
            # The button says which part it belongs to - "Part 2 Field" - so
            # on a phone its ownership is read in the same glance as the press,
            # not looked up from a heading above it. A one-part job has nothing
            # to tell apart, so it says "Field".
            label = LANE_NAMES[phase]
            if multi:
                label = f"Part {part} {label}"
            if recorded:
                label = "Back to " + label
        by_part.setdefault(part, []).append(
            _start_button(label, task_id, part, phase, activity)
        )

    ordered = sorted(by_part, key=lambda p: (p is None, p or 0))
    if not multi:
        merged = [button for part in ordered for button in by_part[part]]
        return [(None, merged)] if merged else []
    # A heading appears only when it has something the buttons cannot say: a
    # tick for each of the part's lanes that is FINISHED. The buttons already
    # name their part, so a heading with nothing to add would say it twice.
    return [
        (f"*Part {part}*{_part_ticks(task, part)}"
         if part is not None and _part_ticks(task, part) else None,
         by_part[part])
        for part in ordered
    ]


def lane_entry_activity(task, part, phase):
    """
    Where a lane is up to, as the activity an assembler returning to it would want.

    Nothing done yet: the setup, because that is what starting a lane is.
    Sheeting already recorded: the sheeting, because that is what coming back
    to it means. Packing has no setup at all.
    """
    if phase == "packing":
        return "production"
    if work_elapsed(task, part, phase, "production"):
        return "production"
    if work_elapsed(task, part, phase, "setup"):
        return "setup"
    return "setup"


def _part_ticks(task, part):
    """
    A tick beside a part for each of its lanes that is finished.

    The one thing the buttons underneath genuinely cannot say. An open lane is
    a button, a started one says "Back to"; but a lane that is DONE has no
    button at all, and bare absence reads the same as a lane the diagram never
    had. A mark on a heading already on screen settles that without describing
    the part back to an assembler who is looking at it.
    """
    return "".join(
        "   ✓ " + LANE_NAMES[phase]
        for phase in ("field_sheeting", "border_sheeting")
        if lane_state(task, part, phase) == "complete"
    )


def other_work_buttons(task, already_offered=()):
    """Every other-work button, flattened - for the callers that only count them."""
    return [button for _, row in other_work_rows(task, already_offered) for button in row]


def lane_needs_details(task, part, phase):
    """
    Whether this lane still has to be described before it can be worked.

    A design and a difficulty are asked for on FIRST ENTRY to a lane, which is
    the moment the assembler is looking at that part. Before parts existed the
    border was described in a form reached by finishing the field; now every
    lane is described the same way, on the way in.
    """
    # Packing has none. It is fetch-box-and-pack: there is no design to name
    # and no difficulty to give it, and asking would be inventing a question.
    #
    # Nor has the job's own opening setup, for the same reason and one more:
    # it is not a lane at all, so lane_of would answer with the border's row
    # and, finding no design on it, send the assembler a form about a part of the
    # diagram when what they pressed was "carry on getting the job ready".
    if phase in ("packing", "job_setup"):
        return False
    lane = lane_of(task, part, phase)
    if not lane.get("present", True):
        return False
    if lane.get("state") == "complete":
        return False
    return not lane.get("design")


def initial_setup_resumable(task):
    """
    Whether the job's own opening setup is still something to carry on with.

    An assembler who starts setting a job up and stops for the day must be able to
    continue THAT setup later. Pressing Pause is not a decision to abandon it.

    But it stays the OPENING setup: once the job has genuinely moved on, it is
    over as a destination and never comes back. So all three must hold, and all
    three are read from the timing the card already reads - the segment ledger
    is the evidence and there is no second record of this:

      the opening setup has time on it   (it really started)
      nothing is being timed             (it really stopped)
      no other work has any time at all  (the job has not moved on)

    That last one is the whole safeguard. It is deliberately not "is the cursor
    still early" or "does the card look like the beginning": a lane with one
    second on it means the assembler left, and leaving is a decision.
    """
    if task.get("working_on"):
        return False
    if not (task.get("job_setup_elapsed") or 0):
        return False
    return not any(
        work_elapsed(task, part, phase, activity)
        for part, phase, activity in every_work(task)
    )


def resume_target(task):
    """
    What Resume means on a paused card: the last thing the assembler was doing.

    Returns (part, phase, activity), or None when there is nothing to resume.

    Read from the ledger rather than assumed, because the last thing they were
    doing is not always the lane the job is on - an assembler who stopped packing
    mid-field is paused on the packing. Two fallbacks follow, in the order a
    assembler would think of them: the lane the cursor is on, and then the first
    unfinished work anywhere on the job. The last of those matters now that a
    job can have several parts - the cursor's own lane may be finished while
    Part 3 has not been touched, and answering None there would leave a paused
    card with nothing to press.

    The opening setup is deliberately not a resume target. It belongs to the
    job rather than to a lane, it is done once at the start, and offering to
    resume it after the sheeting has begun would invite time onto a stage the
    job has left.
    """
    last = task.get("last_work")
    if last and last["phase"] != "job_setup" and lane_open(task, last.get("part"), last["phase"]):
        return last.get("part"), last["phase"], last["activity"]

    cursor_phase = task["current_phase"]
    cursor_part = task.get("current_part")
    if cursor_phase != "completed" and lane_open(task, cursor_part, cursor_phase):
        return cursor_part, cursor_phase, "production"

    for part, phase, activity in every_work(task):
        if activity == "production" and lane_open(task, part, phase):
            return part, phase, activity
    return None


# Slack's ceiling for a header block's text.
HEADER_LIMIT = 150


def header_text(task, suffix=""):
    """
    "T-12  Customer Name", trimmed to something Slack will accept.

    The job number and any suffix are never what gets cut: they are how an assembler
    finds the card. This heads the CLOSED card, where the job is what the card
    is about; a card still being worked heads with the work instead and carries
    the customer's full name in its foot.
    """
    prefix = "T-" + str(task["task_id"]) + "  "
    room = HEADER_LIMIT - len(prefix) - len(suffix)
    name = task["customer_name"] or ""
    if len(name) > room:
        name = name[: max(room - 1, 0)].rstrip() + "…"
    return prefix + name + suffix


def _button(text, action_id, value, style=None, confirm=None):
    button = {
        "type": "button",
        "text": {"type": "plain_text", "text": text},
        "action_id": action_id,
        "value": value,
    }
    if style:
        button["style"] = style
    if confirm:
        button["confirm"] = confirm
    return button


# One action_id per item of work. Slack requires an action_id to be unique
# within its block, and a card now offers every part of unfinished work on the
# job side by side - field sheeting next to border setup next to packing. Ids
# that said only "setup" or "production" collided the moment two lanes were
# offered together, which is the defect that stopped a card rendering at all.
# All of them reach the same handler; the value still carries the job, the lane
# and the activity.
START_ACTION_IDS = {
    ("field_sheeting", "setup"): "trk_start_field_setup",
    ("field_sheeting", "production"): "trk_start_field_production",
    ("border_sheeting", "setup"): "trk_start_border_setup",
    ("border_sheeting", "production"): "trk_start_border_production",
    ("packing", "production"): "trk_start_packing_production",
    # The job's own opening setup. No old card carries this one - the button
    # did not exist before - so there is nothing to stay compatible with.
    ("job_setup", "setup"): "trk_start_job_setup",
}


def _start_button(text, task_id, part, phase, activity, style=None):
    """
    A button that opens an item of work and starts timing it.

    Five action ids, one per lane-and-activity, whatever the job's shape. They
    stay unique because each part's buttons live in their OWN actions block -
    Slack scopes an action_id to its block - so a job drawn as ten parts needs
    no ids that a job drawn as two did not. The value carries which part.
    """
    action_id = START_ACTION_IDS.get(
        (phase, activity),
        "trk_start_setup" if activity == "setup" else "trk_start_production",
    )
    return _button(text, action_id, work_value(task_id, part, phase, activity), style=style)


def _jig_button(task):
    """
    Set jig / Add jig.

    Same action either way - the record is always appended, never overwritten,
    because a jig that was genuinely used stays used. The wording changes
    because "Add" reads as a second one, and the first time there is nothing to
    add to.

    Offered only while the assembler is ON a lane, because the jig belongs to the
    lane being worked and a card that is not on one has nothing to attach it to.
    """
    here = task.get("working_on") or {}
    if here.get("phase") not in ("field_sheeting", "border_sheeting"):
        return None
    lane = lane_of(task, here.get("part"), here["phase"])
    return _button(
        "Add jig" if lane.get("jigs") else "Set jig / template",
        "trk_add_jig",
        work_value(task["task_id"], here.get("part"), here["phase"], "production"),
    )


def _finish_button(task, part, phase):
    """
    The one press that says a lane is done.

    Nothing else on the card finishes anything, and it is deliberately not
    offered until the lane has some sheeting time on it: straight out of setup
    the forward move is to START the sheeting, not to declare it over.
    """
    if not lane_open(task, part, phase):
        return None
    here = task.get("working_on") or {}
    on_it_now = (
        here.get("phase") == phase
        and here.get("part") == part
        and here.get("activity") == "production"
    )
    # Straight out of setup there is nothing to declare finished, so the
    # forward move is to START the sheeting. Once it has started - this
    # instant, not once a minute has accrued - finishing it is a real choice.
    if not on_it_now and work_elapsed(task, part, phase, "production") <= 0:
        return None
    named = work_name(phase, "production", part_label(task, part))
    # The heading already says which part is being worked, so the widest button
    # on the card does not repeat it. The confirmation still names the lane in
    # full, which is where an assembler about to close something for good reads it.
    label = FINISH_LABELS[phase]
    return _button(
        label,
        "trk_complete_task",
        work_value(task["task_id"], part, phase, "production"),
        confirm={
            "title": {"type": "plain_text", "text": "Finished?"},
            # Slack renders a confirmation's text as PLAIN TEXT. Asterisks
            # meant as emphasis are printed, so the assembler read "*field
            # sheeting*" with the asterisks in it.
            "text": {
                "type": "plain_text",
                "text": (
                    "This closes " + lower_name(named) + " for good.\n\n"
                    "Still something to do on it? Use Pause instead - that leaves it "
                    "unfinished and you can come back."
                ),
            },
            "confirm": {"type": "plain_text", "text": "Yes, it is finished"},
            "deny": {"type": "plain_text", "text": "Not yet"},
        },
    )


def job_is_finishable(task):
    """Every lane on every part is done or was never there, and packing too."""
    for part, phase, activity in every_work(task):
        if activity != "production":
            continue
        if lane_open(task, part, phase):
            return False
    return True


def delete_still_applies(task):
    """
    Whether this job could still be one that should never have been entered.

    Delete is that correction, and it is only honest while nothing has been
    made yet. The test is what the job has actually PRODUCED, across every
    part of it - not which lane it happens to sit on, and not whether the
    opening setup has run, because reading the diagram and discovering it is
    the wrong job is exactly when an assembler needs this button.
    """
    for part, phase, activity in every_work(task):
        if activity != "production":
            continue
        if work_elapsed(task, part, phase, "production"):
            return False
        if lane_state(task, part, phase) == "complete":
            return False
    return True


def _delete_button(task_id):
    # The word, not the behaviour. This has never deleted anything: it cancels
    # and keeps the record (see handle_delete), and every other string the
    # assembler meets already says so - the dialog body, its confirm button, and
    # the card they are left with. "Delete" was the last place that did not.
    # The action id is unchanged, so cards already sitting in DMs still work.
    # Cancel is for a job that should not have been entered. It is NOT how an
    # assembler frees themselves to work on another job - Pause does that - so
    # the dialog says only what cancelling does: the job leaves their work
    # list, its record is kept, and this card cannot bring it back.
    return _button(
        "Cancel job",
        "trk_delete_task",
        work_value(task_id),
        style="danger",
        confirm={
            "title": {"type": "plain_text", "text": "Cancel this job?"},
            "text": {
                "type": "plain_text",
                "text": (
                    "The job is cancelled and comes off your work list. Time and history "
                    "already recorded are kept. You cannot resume a cancelled job from this "
                    "card - a supervisor would have to bring it back."
                ),
            },
            "confirm": {"type": "plain_text", "text": "Yes, cancel it"},
            "deny": {"type": "plain_text", "text": "Keep it"},
        },
    )


def card_actions(task):
    """
    What the assembler can do with the work they are HOLDING.

    Only that. Everything else on the job is a row of its own below, one per
    part, so this strip stays short enough to read at a bench. Looking after
    the job itself - correcting its details, taking it off the list - is not
    work, and sits in its own row at the foot rather than among the presses a
    assembler makes while sheeting.

    At most one button here is ever highlighted: the forward move. Where there
    is no forward move there is no highlight.
    """
    task_id = task["task_id"]
    here = task.get("working_on")
    cutting = task.get("cutting_now")
    buttons = []

    if here and here["phase"] == "job_setup":
        # Getting the job ready. The job is not a wizard: no lane is the "next"
        # one, so nothing here is highlighted and no lane is promoted into this
        # strip. Every Field, Border and the packing sit as equal choices in
        # the rows below; the only press that belongs to the work in hand is
        # to stop it.
        buttons.append(_button("Pause setup", "trk_stop_task", work_value(task_id)))
        return buttons

    if here and here["activity"] == "setup":
        # Setting a lane up. The forward move is the sheeting itself, and this
        # is the moment the jig becomes known, so both are on the card. The
        # heading has already said which part, so the button does not.
        part = here.get("part")
        buttons.append(_start_button(
            "Start " + lower_name(work_name(here["phase"], "production")),
            task_id, part, here["phase"], "production", style="primary",
        ))
        buttons.append(_button("Pause setup", "trk_stop_task", work_value(task_id)))
        jig = _jig_button(task)
        if jig:
            buttons.append(jig)
        return buttons

    if here:
        # Working. Pause, measure cutting alongside it, or say the lane is
        # finished. Moving to other work is the rows below, not a form.
        part = here.get("part")
        if cutting:
            buttons.append(_button("Stop cutting", "trk_stop_cutting", work_value(task_id),
                                   style="primary"))
        buttons.append(_button("Pause", "trk_stop_task", work_value(task_id)))
        if not cutting and here["phase"] != "packing" and here["activity"] == "production":
            # Cutting is measured inside THIS part's sheeting, and takes its
            # part from the segment it is inside - so the value names the work
            # it is happening in rather than asking the assembler again.
            buttons.append(_button(
                "Start cutting", "trk_start_cutting",
                work_value(task_id, part, here["phase"], "production"),
            ))
        jig = _jig_button(task)
        if jig:
            buttons.append(jig)
        finish = _finish_button(task, part, here["phase"])
        if finish:
            buttons.append(finish)
        return buttons

    # Nothing is being timed.
    #
    # The job's own opening setup comes first, when it is still live work. It
    # is what the assembler was doing, so it is the forward move; the lanes are
    # somewhere ELSE to go and belong under the rule with the rest of the job.
    # Taking that press is what ends the opening setup as a destination - which
    # is the hierarchy saying so, rather than a line of prose explaining it.
    if initial_setup_resumable(task):
        buttons.append(_start_button(
            "Resume initial setup", task_id, None, "job_setup", "setup",
            style="primary",
        ))
        # No finish check here: if the opening setup is still resumable then no
        # lane has a second on it, so none can be complete and the job cannot
        # be finishable. Adding it would be dead code that risks a second
        # highlighted button.
        return buttons

    resume = resume_target(task)
    if resume:
        part, phase, activity = resume
        touched = (
            work_elapsed(task, part, phase, "setup")
            or work_elapsed(task, part, phase, "production")
        )
        # A press is highlighted here ONLY when there is recorded work to go
        # back to: the assembler stopped part-way through something, and
        # carrying on with it is the forward move. Work nobody has started is
        # a choice, not a continuation, and choices are the rows below - all
        # of them, none coloured in. Highlighting "Start Part 1 Field setup"
        # here made one lane look like the required next step, and the job is
        # not a wizard.
        if touched:
            named = work_name(phase, activity, part_label(task, part))
            buttons.append(_start_button(
                "Resume " + lower_name(named),
                task_id, part, phase, activity, style="primary",
            ))
            if activity == "setup":
                # Setup has been through once, so the sheeting is now a real
                # choice beside carrying on with the setup. Setup is the ONLY
                # way INTO a lane; this is the only place the sheeting is
                # offered before any of it has run.
                buttons.append(_start_button(
                    "Start " + lower_name(work_name(phase, "production", part_label(task, part))),
                    task_id, part, phase, "production",
                ))
    # The one press that ends the job, offered only once there is nothing left
    # unfinished anywhere on it.
    if job_is_finishable(task) and task["current_phase"] != "completed":
        buttons.append(_button("Finish the job", "trk_complete_task",
                               work_value(task_id), style="primary"))
    return buttons


# ---------------------------------------------------------------------------
# The due date
# ---------------------------------------------------------------------------
# One field, meaning DD/MM/YY. Blank means nobody supplied a date - not that
# the job has no deadline, which is not a state this workshop has.

NO_DUE_DATE = "Not set"


DUE_DATE_LABEL = "Due date (DD/MM/YY)"
DUE_DATE_HINT = "e.g. 01/09/26 - or leave blank"
DUE_DATE_ERROR = ("Enter the date as DD/MM/YY, for example 01/09/26. "
                  "If nobody has given you one, leave it blank.")


def read_due_date(typed):
    """
    What the assembler meant by what they typed in the due date box.

    Returns (text to store, error to show them). The text is None when no date
    has been supplied: a blank box, or the "N/A" the old form used to write,
    which is read as the same thing so retyping it cannot mint another one.

    The box says DD/MM/YY and means it. A label that accepts "Friday" is a
    suggestion rather than a promise, and stores a due date nothing can sort by
    or chase. A real date is stored the way the label reads, whatever separator
    was typed and whether the year was given as two digits or four, so every
    card says it the same way.

    The date has to exist: 31/02/26 is refused here rather than accepted and
    then rejected by the database, where the assembler would see nothing useful.

    This governs TYPING ONLY. Rows already holding free text keep it, and are
    read back untouched.
    """
    raw = (typed or "").strip()
    if not raw or raw.upper() == "N/A":
        return None, None
    for separator in ("/", "-", "."):
        parts = raw.split(separator)
        if len(parts) != 3 or not all(p.isdigit() for p in parts):
            continue
        day, month, year = parts
        if len(year) == 2:
            year = "20" + year
        if len(year) != 4:
            break
        try:
            datetime.date(int(year), int(month), int(day))
        except ValueError:
            break
        return "%02d/%02d/%s" % (int(day), int(month), year[2:]), None
    return None, DUE_DATE_ERROR


def due_date_supplied(task):
    """
    The due date the assembler was given, or None when nobody has given them one.

    Every job needs doing as soon as practicable, so there is no such thing as
    a job with no deadline. There are only jobs where a specific calendar date
    is known and jobs where it is not, and a blank box says the second one.

    Rows entered through the old form stored the word "N/A" for that, either as
    the typed text or as a ticked box. It is read as "none supplied" here, at
    the moment it is shown, so an old record reads correctly without anything
    being rewritten. Anything else the assembler typed is theirs and comes back
    untouched - "Friday" and "when the paint arrives" are real answers.
    """
    text = (task.get("due_date") or "").strip()
    if not text or text.upper() == "N/A":
        return None
    return text


def due_date_display(task):
    """The due date as a person reads it, wherever one is shown."""
    return due_date_supplied(task) or NO_DUE_DATE


# ---------------------------------------------------------------------------
# Difficulty
# ---------------------------------------------------------------------------
# A whole number, up to two digits. Every difficulty ever recorded is one, and
# the box has always taken two characters - but it never said so, so an assembler
# who answered in words was told only that they had used too many characters.
# It is deliberately NOT capped at ten: nobody has given us a maximum, and the
# recorded range only tells us what has been typed so far, not what is allowed.

DIFFICULTY_LABEL_FIELD = "Field difficulty"
DIFFICULTY_LABEL_BORDER = "Border difficulty"
DIFFICULTY_HINT = "e.g. 12"
DIFFICULTY_ERROR = "Give the difficulty as a whole number, up to two digits - for example 12."


def read_difficulty(typed):
    """
    Returns (text to store, error to show them). None with no error means the
    box was left blank, which is allowed: not every job has both lanes, and a
    lane the diagram does not have has no difficulty either.
    """
    raw = (typed or "").strip()
    if not raw:
        return None, None
    if not all(character in "0123456789" for character in raw):
        return None, DIFFICULTY_ERROR
    if len(raw) > 2:
        return None, DIFFICULTY_ERROR
    value = int(raw)
    if value < 1:
        return None, DIFFICULTY_ERROR
    return str(value), None


PART_COUNT_ERROR = "How many parts? A whole number, 1 or more."


def read_part_count(typed):
    """
    How many parts the diagram is drawn as.

    A whole number, at least one, at most two digits. THERE IS NO BUSINESS
    MAXIMUM: two digits is the width of the box, not a rule about how many
    parts a job may have, and if a job ever needs more the limit to change is
    that width.

    Returns (count, error). A blank box means one part, because that is what
    the field is pre-filled with and what most jobs are.
    """
    raw = (typed or "").strip()
    if not raw:
        return 1, None
    if not all(character in "0123456789" for character in raw):
        return None, PART_COUNT_ERROR
    if len(raw) > 2:
        return None, PART_COUNT_ERROR
    value = int(raw)
    if value < 1:
        return None, PART_COUNT_ERROR
    return value, None


def _lane_lines(task, part, phase):
    """
    One lane's time, with its parts underneath in the shape they really have.

    Setup and sheeting ADD UP to the lane's total. Cutting does not: it is time
    spent inside the sheeting, already counted in it. Listing both after one
    word - "includes 3m 42s setup, 50s cutting" - put two different
    relationships in one sum, and the arithmetic then did not work: an assembler
    adding those two figures came up short of the lane total and had no way to
    see why. So the parts sit under the lane, and the cutting sits under the
    sheeting it happened in.

    Packing has no setup, so its total IS its one figure and repeating it as a
    part underneath would say nothing.
    """
    here = task.get("working_on") or {}
    on_this_lane = here.get("phase") == phase and here.get("part") == part
    lane = lane_of(task, part, phase)
    if not lane.get("present", True):
        return ["*" + LANE_NAMES[phase] + "*  not on this part"]
    setup = work_elapsed(task, part, phase, "setup")
    production = work_elapsed(task, part, phase, "production")
    if not setup and not production and not on_this_lane:
        return []

    lines = ["*" + LANE_NAMES[phase] + "*  " + database.format_duration(setup + production)]
    if phase == "packing":
        return lines

    if setup or (on_this_lane and here.get("activity") == "setup"):
        lines.append("•  Setup  " + database.format_duration(setup))
    if production or (on_this_lane and here.get("activity") == "production"):
        lines.append("•  Sheeting  " + database.format_duration(production))
        cutting = lane.get("cutting_elapsed") or 0
        if cutting:
            # Cutting is written under the sheeting it happened inside, never
            # as a line of its own, because it is time already counted. An en
            # dash rather than a hollow bullet: both read as a level below, and
            # this one still prints where the card text is rendered outside
            # Slack.
            lines.append("     –  of which cutting  " + database.format_duration(cutting))
    if lane.get("design") or lane.get("jigs"):
        described = [bit for bit in (
            lane.get("design"),
            ("difficulty " + lane["difficulty"]) if lane.get("difficulty") else None,
            ("jig " + lane["jigs"]) if lane.get("jigs") else None,
        ) if bit]
        lines.append("     " + "  ·  ".join(described))
    return lines


def _time_lines(task, total_label="Total job time"):
    """
    What has been recorded, part by part.

    A lane appears once there is something to say about it, so an early card is
    short and a late one is complete. The parts are headed only when there is
    more than one - on a single-part job "Part 1" is a heading with nothing to
    distinguish it from.
    """
    multi = (task.get("part_count") or 1) > 1
    rows = []

    setup = task.get("job_setup_elapsed") or 0
    here = task.get("working_on") or {}
    if setup or here.get("phase") == "job_setup":
        # The job's own preparation, above the parts, because that is what it
        # is: work on the job before any one part of it.
        rows.append("*Initial setup*  " + database.format_duration(setup))

    for row in task.get("parts") or []:
        part = row.get("part")
        lane_rows = (_lane_lines(task, part, "field_sheeting")
                     + _lane_lines(task, part, "border_sheeting"))
        if not lane_rows:
            continue
        if multi:
            rows.append("*Part " + str(part) + "*")
        rows += lane_rows

    rows += _lane_lines(task, None, "packing")

    # Nothing worked yet: the status line has already said what the assembler is on
    # and how long for, and repeating it under a heading is three noughts and no
    # information.
    if not rows or not task["total_elapsed"]:
        return []
    return ["*Time recorded*"] + rows + [
        "*" + total_label + "*  " + database.format_duration(task["total_elapsed"])
    ]


def _headline(task):
    """
    The one thing the card is about: the work that is being timed right now.

    Slack's header is the only block with its own type size, so it carries the
    state and nothing else - the job number and the customer have all day to be
    read and sit in the foot. Cutting takes the heading while it runs because
    that is what the assembler is doing, and it names the sheeting it is inside so
    the time is never in doubt.
    """
    here = task.get("working_on")
    if not here:
        return MARK_PAUSED + "  Paused - nothing is being timed"
    cutting = task.get("cutting_now")
    if cutting:
        inside = work_name(cutting["parent_phase"], "production",
                           part_label(task, cutting.get("part")))
        return MARK_CUTTING + "  " + inside + " - cutting now"
    if here["phase"] == "job_setup":
        return MARK_RUNNING + "  Initial setup - running"
    named = work_name(here["phase"], here["activity"], part_label(task, here.get("part")))
    return MARK_RUNNING + "  " + named + " - running"


def _facts_line(task):
    """
    One grey line under the heading, or nothing at all.

    The only thing allowed between the work and the buttons, and it earns the
    place by being about the work in hand: which design, how hard, which jig -
    what an assembler checks against the diagram without leaving the bench. Where
    there is nothing true to put here the line is not drawn; nothing is invented
    to fill it.

    A paused card uses it for the one figure that is finished and therefore
    safe to print.
    """
    here = task.get("working_on")
    if not here:
        last = task.get("last_work")
        if not last:
            return None
        if last["phase"] == "job_setup":
            name = "Initial setup"
            recorded = task.get("job_setup_elapsed") or 0
        else:
            name = work_name(last["phase"], last["activity"], part_label(task, last.get("part")))
            recorded = work_elapsed(task, last.get("part"), last["phase"], last["activity"])
        line = "Last on " + lower_name(name)
        # A figure of nought says less than no figure: it reads as a timer that
        # did not work, when the assembler simply moved on within the second.
        if recorded:
            line += "  ·  " + database.format_duration(recorded) + " recorded"
        return line

    cutting = task.get("cutting_now")
    phase = cutting["parent_phase"] if cutting else here["phase"]
    part = cutting.get("part") if cutting else here.get("part")
    if phase not in ("field_sheeting", "border_sheeting"):
        return None
    lane = lane_of(task, part, phase)
    described = [bit for bit in (
        lane.get("design"),
        ("difficulty " + lane["difficulty"]) if lane.get("difficulty") else None,
        ("jig " + lane["jigs"]) if lane.get("jigs") else None,
    ) if bit]
    return "  ·  ".join(described) or None


def _admin_actions(task):
    """
    Looking after the job rather than working it: correct it, or take it off
    the list.

    Its own row at the foot, because an assembler reaching for Pause should not find
    Delete beside it. Offered in exactly the states that offered it before - a
    card in the middle of sheeting has never carried either, and moving a strip
    around is no reason to start.
    """
    here = task.get("working_on") or {}
    if here and here.get("phase") != "job_setup" and here.get("activity") != "setup":
        return []
    buttons = [_button("Edit details", "trk_edit_task", work_value(task["task_id"]))]
    if delete_still_applies(task):
        buttons.append(_delete_button(task["task_id"]))
    return buttons


def _parts_finished(task):
    """
    How much of a multi-part job is behind the assembler, as (done, total).

    None on a job drawn as one part, where "0 of 1 parts finished" says only
    that the job is not finished, which the card has already said - and None
    again until the first part is actually done, for the same reason. A part
    is done when neither of its lanes is still open; a lane the diagram never
    had was never work.
    """
    rows = task.get("parts") or []
    if len(rows) < 2:
        return None
    done = sum(
        1 for row in rows
        if not lane_open(task, row.get("part"), "field_sheeting")
        and not lane_open(task, row.get("part"), "border_sheeting")
    )
    return (done, len(rows)) if done else None


def _foot_lines(task):
    """
    Which job this is, in grey, at the bottom. Two lines, read once.

    The customer's name goes in FULL and unconditionally. The heading is given
    over to the work now, so this is the only place the card carries the name at
    all, and the name most likely to matter is the long one a heading would have
    cut.
    """
    here = task.get("working_on") or {}
    first = "  ·  ".join(str(bit) for bit in (
        "T-" + str(task["task_id"]),
        task["customer_name"],
        task["task_description"],
    ) if bit)

    second = [str(task["invoice_number"]), "due " + due_date_display(task)]
    counted = _parts_finished(task)
    if counted:
        second.append(f"{counted[0]} of {counted[1]} parts finished")
    if not here and task["total_elapsed"]:
        # Only where nothing is accruing, so the figure is final rather than a
        # snapshot that went stale the moment it was drawn.
        second.append(database.format_duration(task["total_elapsed"]) + " on this job")
    return [first, "  ·  ".join(second)]


def job_card(task, note=None):
    """
    The whole card. Returns (fallback text, blocks).

    A WORKING INTERFACE, in the order an assembler at a bench asks for it: what am I
    on, what can I do with it, what else could I move to. The job's own details
    are true all day and are read once, so they go grey at the foot.

    WHILE A TIMER IS RUNNING THE CARD PRINTS NO DURATION ANYWHERE. Slack does
    not tick and a card is only redrawn by a press, so a figure beside running
    work is already wrong when it is drawn - the card read "0s" beside work it
    said in the same breath was running. Figures appear only where nothing is
    accruing and they are final. None of the detail is lost: the full breakdown
    is on the closed card, in the ledger and in the export, which is where
    somebody reading a job back actually looks.

    `note` is a single grey line BELOW the buttons when something just happened
    that the card alone would not explain. Below, because a confirmation must
    not push the work the assembler is holding down the card.
    """
    task_id = task["task_id"]
    blocks = [{
        "type": "header",
        "text": {"type": "plain_text", "text": _headline(task), "emoji": True},
    }]

    facts = _facts_line(task)
    if facts:
        blocks.append({"type": "context", "elements": [{"type": "mrkdwn", "text": facts}]})

    actions = card_actions(task)
    if actions:
        blocks.append({
            "type": "actions",
            "block_id": "task_actions_" + str(task_id),
            "elements": actions,
        })

    if note:
        blocks.append({"type": "context", "elements": [{"type": "mrkdwn", "text": note}]})

    # The rest of the job, one row per part. An assembler is not walking a wizard:
    # while the job is unfinished they can move between whatever work it
    # actually contains, and the card says so rather than hiding it behind a
    # form. What moving DOES is not explained here every time - the buttons
    # under a heading are the explanation, and an assembler who presses one finds out
    # once rather than reading it on every render.
    #
    # Each part gets its OWN actions block. That is not only for reading:
    # Slack scopes an action_id to its block, and Part 1's field sheeting and
    # Part 2's field sheeting share one - side by side in a single block the
    # card would not render at all.
    offered = {b.get("value") for b in actions
               if (b.get("action_id") or "").startswith("trk_start_")}
    rows = other_work_rows(task, offered)
    if rows:
        # "Other" only while a lane or the packing is actually being worked;
        # during the job's own setup, or paused, nothing below is "other" -
        # it is simply the work the job has.
        here = task.get("working_on") or {}
        on_a_lane = bool(here) and here.get("phase") != "job_setup"
        blocks.append({"type": "divider"})
        blocks.append({
            "type": "context",
            "elements": [{"type": "mrkdwn",
                          "text": "*Other work on this job*" if on_a_lane
                          else "*Work on this job*"}],
        })
        for index, (heading, buttons) in enumerate(rows, start=1):
            if heading:
                blocks.append({
                    "type": "context",
                    "elements": [{"type": "mrkdwn", "text": heading}],
                })
            blocks.append({
                "type": "actions",
                "block_id": f"task_elsewhere_{index}_{task_id}",
                "elements": buttons,
            })

    admin = _admin_actions(task)
    if admin:
        blocks.append({
            "type": "actions",
            "block_id": "task_admin_" + str(task_id),
            "elements": admin,
        })

    blocks.append({
        "type": "context",
        "elements": [{"type": "mrkdwn", "text": "\n".join(_foot_lines(task))}],
    })

    here = task.get("working_on") or {}
    if not here:
        summary = "T-" + str(task_id) + ": paused"
    elif here["phase"] == "job_setup":
        summary = "T-" + str(task_id) + ": initial setup"
    else:
        summary = "T-" + str(task_id) + ": " + work_name(
            here["phase"], here["activity"], part_label(task, here.get("part")))
    return summary, blocks


def update_card(client, task, channel_id, note=None):
    """
    Rewrite the job's card where it already is - or post one, if the job has
    somehow ended up without a card to rewrite.
    """
    if not task.get("message_ts"):
        repost_card(client, task, channel_id, note=note)
        return
    text, blocks = job_card(task, note=note)
    client.chat_update(
        channel=channel_id,
        ts=task["message_ts"],
        text=text,
        blocks=blocks,
    )


def repost_card(client, task, channel_id, note=None):
    """
    Replace the card with a fresh one at the bottom of the DM.

    Used at the points where the job genuinely moves on - a lane finished, the
    border decided - because by then the old card has usually scrolled away
    behind the modal that was just filled in, and an assembler should not have to go
    looking for the job they are working on.
    """
    text, blocks = job_card(task, note=note)
    if task.get("message_ts"):
        try:
            client.chat_delete(channel=channel_id, ts=task["message_ts"])
        except SlackApiError:
            # Already gone is the outcome this wanted anyway.
            pass
    result = client.chat_postMessage(channel=channel_id, text=text, blocks=blocks)
    database.update_message_ts(task["task_id"], result["channel"], result["ts"])
    return result


def resolve_job(client, body, task_id, user_id, channel_id):
    """
    The three things every button checks before it does anything: the job is
    still there, it belongs to this assembler, and it is not already finished.

    Returns the job, or None having already said why.
    """
    task = database.get_task(task_id)
    if task is None:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="That job is not there any more - it may have been deleted.",
        )
        return None
    if task["user_id"] != user_id:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="This is <@" + task["user_id"] + ">'s job, so only they can change it.",
        )
        return None
    if task["current_phase"] == "completed":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="T-" + str(task_id) + " is finished, so nothing more can be recorded on it.",
        )
        return None
    return task


def busy_elsewhere_text(active):
    """
    The one sentence an assembler gets when they try to start work while a
    DIFFERENT job of theirs is being timed: which job it is, and that Pause on
    that job is what frees them. Nothing is switched or paused on their behalf.

    `active` is the job being timed, as get_active_task returns it - read
    fresh, so the number is the job that is timing NOW rather than the one a
    stale card last knew about. None means it stopped in the meantime.
    """
    if active:
        return (
            "You're already working on T-" + str(active["task_id"]) + " "
            + active["customer_name"] + ". Pause that job before starting this one."
        )
    return "You're already working on another job. Pause that one before starting this."


def busy_elsewhere(client, task, user_id, channel_id):
    """
    Is this assembler timing a DIFFERENT job right now? Then say so and answer
    True, so the press stops here - before a form opens that would be filled
    in for nothing. LMSA refuses the start itself either way; this only spares
    the assembler the form.
    """
    active = database.get_active_task(user_id)
    if not active or active["task_id"] == task["task_id"]:
        return False
    client.chat_postEphemeral(channel=channel_id, user=user_id, text=busy_elsewhere_text(active))
    return True


def refusal_text(reason, task, phase=None):
    """
    A refusal, in workshop words: what happened, and what to do about it.

    Every one of these is something a real assembler can cause by pressing a
    button, usually from a card that has gone stale in another window. So the
    answer says what the job is actually doing now and what to press instead -
    never a reason code, and never nothing at all.
    """
    if reason == "another_job_running":
        return busy_elsewhere_text(database.get_active_task(task["user_id"]))
    lane = lower_name(work_name(phase or task["current_phase"], "production"))
    here = task.get("working_on")
    doing = lower_name(work_name(here["phase"], here["activity"])) if here else None
    texts = {
        "already_running": "You are already on that, so nothing has changed.",
        "another_phase_running": (
            "The " + (doing or "other") + " timer is running on this job. Pause it first, "
            "then try again."
        ),
        "other_activity_running": (
            "You are on " + (doing or "other work") + " on this job. Pick it from *Other work "
            "on this job* on the card - that pauses what you are doing rather than finishing it."
        ),
        "phase_already_complete": (
            "The " + lane + " is already finished on this job, so it cannot be started again."
        ),
        "phase_already_skipped": (
            "This job is marked as having no border, so there is no border to work on. If that "
            "was wrong, use *Border after all* on the packing card."
        ),
        "phase_has_no_setup": "Packing has no setup step - there is nothing to prepare.",
        "cutting_needs_production_work": (
            "Start the sheeting first. Cutting is recorded as part of the field or border work "
            "it happens during, so there has to be some running."
        ),
        "not_running": (
            "Nothing is being timed on this job at the moment - your card will show "
            "what this job is up to and what you can start."
        ),
        "already_cutting": "The cutting is already being timed.",
        "not_cutting": (
            "No cutting is being timed on this job at the moment - your card will show "
            "what this job is up to and what you can start."
        ),
        "job_not_open": "This job is no longer open, so nothing has been changed.",
    }
    return texts.get(reason, "That could not be done, so nothing has been changed.")


# ---------------------------------------------------------------------------
# Commands, and the handlers behind every button
# ---------------------------------------------------------------------------
# From here down, every function is registered with Slack. The middleware
# runs first on each delivery and notes which delivery it is, so a click that
# arrives twice is recognised as the same action.

@app.middleware
def track_slack_delivery(body, next):
    # Note which Slack delivery is being handled, so a redelivery of the same
    # click is recognised as the same action rather than applied twice. Scoped
    # to this handler and cleared afterwards, so nothing carries over into the
    # next request. Changes no behaviour an assembler can see.
    with database.slack_request(body):
        return next()

# A liveness check: it answers, so the bot is connected and its commands are
# registered. Nothing in the workflow uses it.

@app.command ("/hello")
def hello_command(ack, body, say):

    ack()
    user_id = body["user_id"]
    say(f"Hi there, <@{user_id}>! I'm ready to track your projects.")

# /track opens the intake form. /track export sends the spreadsheet instead.

@app.command("/track")
def track_command(ack, body, client):
    ack()
    user_id = body ["user_id"]

    subcommand = body.get ("text", "").strip().lower()
    if subcommand == "export":
        handle_export(body,client)
        return

    # One timer at a time per person. An assembler may hold several unfinished
    # jobs - each paused one keeps its card - but creating a job starts its
    # setup clock, so an assembler who is timing another job right now is told
    # which one, and that Pause on it is what frees them. Nothing is paused
    # on their behalf.
    active = database.get_active_task(user_id)
    if active:
        client.chat_postEphemeral(
            channel=body["channel_id"],
            user=user_id,
            text=busy_elsewhere_text(active),
        )
        return
    # The first of the two intake forms. private_metadata carries the channel
    # the command came from, so the job can be announced back to the same room.
    client.views_open(
        trigger_id=body["trigger_id"],
        view={
            "type": "modal",
            "callback_id": "trk_track_step_1",  # ID used to catch the submission
            "title": {"type": "plain_text", "text": "New job - 1 of 2"},
            "private_metadata":body["channel_id"],
            "submit": {"type": "plain_text", "text": "Next"},
            "close": {"type": "plain_text", "text": "Cancel"},
            "blocks": [
                {
                    "type": "input",
                    "block_id": "customer_block",
                    "label": {"type": "plain_text", "text": "Customer name"},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "customer_name"
                    }
                },
                {
                    "type": "input",
                    "block_id": "invoice_block",
                    "label": {"type": "plain_text", "text": "Invoice / Pro Forma number"},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "invoice_num"
                    }
                },
                {
                    "type": "input",
                    "block_id": "task_block",
                    "label": {"type": "plain_text", "text": "Job description"},
                    "element": {
                        "type": "plain_text_input",
                        "multiline": True,
                        "action_id": "task_desc"
                    }
                },
                {
                    "type": "input",
                    "block_id": "date_block",
                    "optional": True,
                    "label": {"type": "plain_text", "text": DUE_DATE_LABEL},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "due_date",
                        "placeholder": {"type": "plain_text", "text": DUE_DATE_HINT}
                    }
                },
                {
                    "type": "input",
                    "block_id": "parts_block",
                    "label": {"type": "plain_text", "text": "Number of parts"},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "part_count",
                        "initial_value": "1",
                        "max_length": 2,
                    },
                    "hint": {"type": "plain_text",
                             "text": "How many parts the diagram is drawn as."},
                }
            ]
        }
    )

# ---------------------------------------------------------------------------
# The spreadsheet export
# ---------------------------------------------------------------------------

def resolve_existing_dm(client, user_id):
    # Find the bot's existing DM conversation with a user, paging through the
    # full list rather than trusting the first page.
    #
    # conversations_open would mint the conversation, but it needs the im:write
    # scope, which the LMSA Slack app does not hold. Listing the bot's own IM
    # conversations needs only im:read. Returns None when no DM exists yet --
    # files_upload_v2 rejects a raw user id, so the caller must handle that
    # rather than passing user_id through.
    cursor = None
    while True:
        resp = client.users_conversations(types="im", limit=200, cursor=cursor)
        for conversation in resp["channels"]:
            if conversation.get("user") == user_id:
                return conversation["id"]
        cursor = (resp.get("response_metadata") or {}).get("next_cursor") or None
        if not cursor:
            return None

def handle_export(body, client):
    user_id = body["user_id"]
    channel_id = body ["channel_id"]
    
    tasks = database.get_completed_tasks()
    
    user_name_cache = {}
    
    def get_user_name(slack_user_id):
        if slack_user_id in user_name_cache:
            return user_name_cache[slack_user_id]

        try:
            response = client.users_info(user=slack_user_id)
            profile = response.get("user", {}).get("profile", {})
            name = profile.get("display_name") or profile.get("real_name") or slack_user_id
        except SlackApiError:
            name = slack_user_id

        user_name_cache[slack_user_id] = name
        return name
    
    if not tasks:
        client.chat_postEphemeral(
            channel = channel_id,
            user = user_id,
            text = "No completed jobs found to export yet."
        )
        return
    
    # Building the Excel Spreadsheet
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if ws is None:
        ws = wb.create_sheet(title="Completed Jobs")
    else:
        ws.title = "Completed Jobs"
    
    # Headers
    
    headers = [
        "Task ID",
        "Customer",
        "Invoice Number",
        "Task Description",
        "Due Date",
        "Field Design",
        "Field Difficulty",
        "Field Jig Size(s)",
        # The lane total, then how it was made up. Setup and sheeting ADD UP to
        # the lane total; cutting is time already inside the sheeting, so it is
        # a breakdown of it and must never be added on.
        "Field Time",
        "Field Setup Time",
        "Field Sheeting Time",
        "Field Cutting (within sheeting)",
        "Border Design",
        "Border Difficulty",
        "Border Jig Size(s)",
        "Border Time",
        "Border Setup Time",
        "Border Sheeting Time",
        "Border Cutting (within sheeting)",
        "Packing Time",
        "Total Time",
        "General Notes",
        "Issues Encountered",
        "Completed By",
        "Date Created",
    ]
    ws.append(headers)
    
    # Styling the headers

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor ="2C3E50")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Data Rows
    for task in tasks:
        field_elapsed = task["field_elapsed"] or 0
        border_elapsed = task["border_elapsed"] or 0
        packing_elapsed = task["packing_elapsed"] or 0
        total_elapsed = field_elapsed + border_elapsed + packing_elapsed

        ws.append([
            f"T-{task['task_id']}",
            task["customer_name"],
            task["invoice_number"],
            task["task_description"],
            due_date_display(task),
            task["field_design"] or "-",
            task["difficulty"],
            # Several jigs show as one readable cell, e.g. "49.6 / 50"
            task["field_jigs"] or "-",
            database.format_elapsed(field_elapsed),
            database.format_elapsed(task["field_setup_elapsed"] or 0),
            database.format_elapsed(task["field_production_elapsed"] or 0),
            database.format_elapsed(task["field_cutting_elapsed"] or 0),
            "No Border" if task.get("border_skipped") else (task["border_design"] or "-"),
            "-" if task.get("border_skipped") else (task["border_difficulty"] or "-"),
            "-" if task.get("border_skipped") else (task["border_jigs"] or "-"),
            border_time_display(task),
            "-" if task.get("border_skipped") else database.format_elapsed(task["border_setup_elapsed"] or 0),
            "-" if task.get("border_skipped") else database.format_elapsed(task["border_production_elapsed"] or 0),
            "-" if task.get("border_skipped") else database.format_elapsed(task["border_cutting_elapsed"] or 0),
            database.format_elapsed(packing_elapsed),
            database.format_elapsed(total_elapsed),
            task["general_notes"] or "None",
            task["issues_encountered"] or "None",
            get_user_name(task["user_id"]),
            task["created_at"],
        ])
        
    #Auto-sizing columns, so the text can fit
    for column_cells in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 4,50)
        
    # Writing to in-memory bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    dm_channel_id = resolve_existing_dm(client, user_id)
    if not dm_channel_id:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=("Unable to send the export because you do not have a direct message "
                  "conversation with the bot yet. Start a job with `/track` first, then "
                  "run the export again.")
        )
        return

    #Upload Files to user's DM
    try:
        client.files_upload_v2(
            channel = dm_channel_id,
            file=buffer.getvalue(),
            filename = "trackbot_jobs_export.xlsx",
            title = f"Trackbot Export - {len(tasks)} Completed Job(s)"
        )
    except SlackApiError as err:
        error_code = err.response.get("error") if err.response is not None else "unknown_error"
        if error_code == "missing_scope":
            client.chat_postEphemeral(
                channel=channel_id,
                user=user_id,
                text=("Unable to upload export because the Slack app is missing the required "
                      "`files:write` scope. Please update the app scopes and reinstall the app.")
            )
            return
        raise
        
        #Confirmation of in the channel of the command working
        
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=f"Export ready! Check your DMs - {len(tasks)} completed job(s) exported to Excel."
        )

# ---------------------------------------------------------------------------
# Job intake
# ---------------------------------------------------------------------------
# Two forms, then the job exists. Submitting the second one is the handover
# into workshop work: it creates the job and starts its setup in the same
# transaction, which is why the card that follows has no Start button.

@app.view("trk_track_step_1")
def handle_step_1(ack,body,client,):
    vals = body["view"]["state"]["values"]
    channel_id = body["view"]["private_metadata"]

    customer_name = vals["customer_block"]["customer_name"]["value"]
    invoice_number = vals["invoice_block"]["invoice_num"]["value"]
    task_description = vals["task_block"]["task_desc"]["value"]
    # An empty box is carried as nothing at all, not as the word "N/A". Nobody
    # has given this assembler a date yet; that is not a job with no deadline.
    due_date, due_date_error = read_due_date(vals["date_block"]["due_date"]["value"])
    if due_date_error:
        # Sent back to the box it belongs to, so the assembler reads the message
        # under the date rather than losing the whole form.
        ack(response_action="errors", errors={"date_block": due_date_error})
        return

    part_count, part_count_error = read_part_count(
        _typed(vals, "parts_block", "part_count"))
    if part_count_error:
        ack(response_action="errors", errors={"parts_block": part_count_error})
        return

    # private_metadata is the only way to carry these across to the pushed
    # form, which arrives as a separate submission.
    step1_data = {
        "channel_id": channel_id,
        "customer_name": customer_name,
        "invoice_number": invoice_number,
        "task_description": task_description,
        "due_date": due_date,
        "part_count": part_count,
    }

    ack(response_action="push", view=step_2_view(step1_data))


def step_2_view(step1_data):
    """
    Screen 2: WHAT WORK EXISTS, and nothing else.

    One section per part, each with a Field and a Border tick. That is the
    whole form. It deliberately does not ask for designs, difficulty, jig or
    cutting: the assembler filling this in has just been handed the diagram and is
    establishing the shape of the job. The details are asked for when they
    first enter that part's lane, which is the moment they are looking at it.

    Checkboxes rather than a text box per lane, because the question is yes or
    no and a tick answers it without the assembler having to know that leaving a
    box empty is how you say "there isn't one".

    Every part needs at least one lane, and the FORM enforces it: the block is
    required, so Slack refuses the submission before it is sent. A part with
    neither lane is a part that exists and can never be worked, and that is
    checked again in the handler and once more by the database - three places,
    because an assembler who already had this form open when it changed still submits
    the shape they were given. None of that is explained on the screen. A rule
    the form will not let you break does not need a paragraph telling you not to
    break it.
    """
    count = step1_data["part_count"]
    blocks = [{
        "type": "section",
        "text": {"type": "mrkdwn", "text": "*What work is on each part?*"},
    }]
    for number in range(1, count + 1):
        # One tick pair per part, headed by the part's own name - the label IS
        # the heading, so the form never says "Part 2" twice. On a one-part job
        # there is nothing to tell it apart from, so it is simply "Work".
        #
        # REQUIRED, and that is the whole of the rule. Left optional, Slack
        # printed "(optional)" after the label - the form said the opposite of
        # what the workshop means, and a sentence above it was being used to
        # argue with its own control.
        blocks.append({
            "type": "input",
            "block_id": f"part_{number}",
            "label": {"type": "plain_text",
                      "text": "Work" if count == 1 else f"Part {number}"},
            "element": {
                "type": "checkboxes",
                "action_id": "lanes",
                "options": [
                    {"text": {"type": "plain_text", "text": "Field"}, "value": "field"},
                    {"text": {"type": "plain_text", "text": "Border"}, "value": "border"},
                ],
            },
        })
    blocks.append({
        "type": "context",
        "elements": [{"type": "mrkdwn",
                      "text": "Creating the job starts recording Initial setup time "
                              "and sends your work card to you in a DM."}],
    })
    return {
        "type": "modal",
        "callback_id": "trk_track_step_2",
        "title": {"type": "plain_text", "text": "New job - 2 of 2"},
        "submit": {"type": "plain_text", "text": "Create the job"},
        "close": {"type": "plain_text", "text": "Back"},
        "private_metadata": json.dumps(step1_data),
        "blocks": blocks,
    }


def read_lanes(values, number):
    """Which lanes were ticked for one part, as a set."""
    block = values.get(f"part_{number}") or {}
    chosen = (block.get("lanes") or {}).get("selected_options") or []
    return {option.get("value") for option in chosen}


def _typed(values, block_id, action_id):
    """
    What the assembler typed into one box, or None if that box was not on the form
    they submitted.

    A modal an assembler already had open when the form changed still submits the
    blocks it was built with. Reading those defensively means an older form
    lands as the job it describes - a field job with no border on it - rather
    than raising on a block that was never there.
    """
    block = values.get(block_id) or {}
    return (block.get(action_id) or {}).get("value")


@app.view("trk_track_step_2")
def handle_step_2(ack, body, client):
    user_id = body["user"]["id"]
    vals = body["view"]["state"]["values"]

    prev_data = json.loads(body["view"]["private_metadata"])
    team_channel_id = prev_data["channel_id"]
    count = prev_data.get("part_count") or 1

    # Every part needs work on it. A part with neither lane is a part that
    # exists and can never be worked, and the error goes against the part it
    # is about rather than the top of the form.
    parts = []
    for number in range(1, count + 1):
        lanes = read_lanes(vals, number)
        if not lanes:
            ack(response_action="errors", errors={
                f"part_{number}": "Tick Field, Border, or both.",
            })
            return
        parts.append({"field": "field" in lanes, "border": "border" in lanes})

    ack(response_action="clear")

    try:
        task_id = database.create_task(
            user_id=user_id,
            channel_id=team_channel_id,
            customer_name=prev_data["customer_name"],
            invoice_number=prev_data["invoice_number"],
            task_description=prev_data["task_description"],
            due_date=prev_data["due_date"],
            parts=parts,
        )
    except database.TrackerRefused as refusal:
        # /track checked this before the form opened, but the assembler may have
        # resumed another job from its card while the form was up. Creating
        # would start a second timer, so LMSA said no and nothing was made.
        if refusal.reason != "another_job_running":
            raise
        client.chat_postEphemeral(
            channel=team_channel_id,
            user=user_id,
            text=busy_elsewhere_text(database.get_active_task(user_id)),
        )
        return

    # Submitting this form is the handover into the workshop: the assembler has the
    # job and is already getting it ready. So the setup timer is running by the
    # time the card appears, and there is no "Start" button - there is nothing
    # left to start. That setup is the JOB's, not the first part's.
    task = database.get_task(task_id)

    # chat_postMessage accepts a user id and resolves the DM itself, returning
    # the real D... conversation id in result["channel"]. conversations_open
    # would need the im:write scope, which the LMSA Slack app does not hold.
    text, blocks = job_card(task)
    result = client.chat_postMessage(channel=user_id, text=text, blocks=blocks)

    # Saving the timestamp
    database.update_message_ts(task_id, result["channel"], result["ts"])

    # THE PUBLIC CHANNEL CARRIES NO TIMING. Who is making what, and later that
    # it is finished. Everything about how long it took lives in the DM card,
    # the history and the export, where the person reading it is the person it
    # is about.
    #
    # A person, then what they are doing - the room is reading about a colleague
    # picking up a job, not a row being written. Only starting and finishing:
    # moving between a field and a border is the assembler's own business and the
    # channel never hears about it.
    client.chat_postMessage(
        channel=team_channel_id,
        text=f"{MARK_RUNNING} <@{user_id}> has started T-{task_id} {task['customer_name']}",
    )


# ---------------------------------------------------------------------------
# Working the job: starting, pausing, cutting, switching
# ---------------------------------------------------------------------------
# Exactly one item of work accrues at a time. Pause means "not working this
# job for now"; Switch work means "still on this job, on something else".
# Neither finishes anything. Cutting is measured inside sheeting and leaves
# the sheeting timer running.

@app.action("trk_start_field_setup")
@app.action("trk_start_field_production")
@app.action("trk_start_border_setup")
@app.action("trk_start_border_production")
@app.action("trk_start_job_setup")
@app.action("trk_start_packing_production")
# The two ids cards used before every item of work had its own. A card posted
# under the older scheme is still sitting in an assembler's DM and still clickable.
@app.action("trk_start_setup")
@app.action("trk_start_production")
def handle_start(ack, body, client):
    """
    Move onto an item of work and start timing it.

    One handler behind every button that does that: Start field sheeting,
    Resume, Start Part 2 border setup and Back to Part 1 field sheeting. The
    value carries which part, which lane and which activity; a button posted
    by an earlier version of the tracker carries less, and what it leaves out
    resolves to where the job's cursor is - the only thing such a card could
    have meant.

    FIRST ENTRY TO A LANE ASKS WHAT IT IS. A lane nobody has described yet gets
    its form here rather than being started blind: the assembler is looking at that
    part of the diagram at exactly this moment, which is why the question is
    asked now and not at intake, and not when some other lane finished.
    """
    ack()
    task_id, part, phase, activity = read_work_value(body["actions"][0]["value"])
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return
    # A card can be pressed while a different job is being timed - the assembler
    # found an old card in their DM. Refused before any form opens: the timer
    # they are running has to be paused first, and nothing here does that for
    # them.
    if busy_elsewhere(client, task, user_id, channel_id):
        return

    phase = phase or task["current_phase"]
    activity = activity or "production"
    if part is None and phase in ("field_sheeting", "border_sheeting"):
        part = task.get("current_part")

    if lane_needs_details(task, part, phase):
        client.views_open(
            trigger_id=body["trigger_id"],
            view=lane_details_view(task, part, phase, activity, channel_id),
        )
        return

    outcome = database.start_work(task_id, phase, activity, part=part)
    if outcome != "started":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=refusal_text(outcome, task, phase),
        )
        return

    update_card(client, database.get_task(task_id), channel_id)


def lane_details_view(task, part, phase, activity, channel_id):
    """
    What this lane is, asked on the way into it.

    Two questions, and only two: the design and the difficulty, because a lane
    with neither cannot be read back later. Both are known from the diagram the
    assembler is holding as they answer.

    THE JIG IS NOT ASKED HERE. Finding and testing it IS the setup, so at the
    moment this form opens the assembler frequently does not know it yet, and a box
    they cannot fill is a question that teaches them to skip questions. It is
    recorded from the work card instead, with "Set jig / template", at the point
    it becomes known - which is where it was always genuinely established.

    Saving starts the work the assembler pressed for. That is the whole point of
    asking here: the form is on the way to the bench, not a detour from it.
    """
    # Name it for what the press starts, not for the lane's sheeting.
    #
    # This was hardcoded to "production", which was right while an untouched
    # lane could be entered straight at its sheeting. It cannot any more - the
    # only way into a lane nobody has touched is its setup - so an assembler who
    # pressed "Start field setup" met a form headed "Field sheeting" and then
    # found their setup running. The form is the same two questions either way;
    # only its heading has to agree with the button that opened it.
    named = work_name(phase, activity, part_label(task, part))
    lane = lane_of(task, part, phase)
    which = "field" if phase == "field_sheeting" else "border"
    return {
        "type": "modal",
        "callback_id": "trk_lane_details",
        # The title carries the identity - "Part 2 Border setup" - and the body
        # goes straight to the two questions. Slack allows 24 characters in a
        # title and silently refuses the view above that.
        "title": {"type": "plain_text", "text": named[:24]},
        "submit": {"type": "plain_text", "text": "Save and start"},
        "close": {"type": "plain_text", "text": "Cancel"},
        "private_metadata": json.dumps({
            "task_id": task["task_id"],
            "part": part,
            "phase": phase,
            "activity": activity,
            "channel_id": channel_id,
        }),
        "blocks": [
            {
                "type": "input",
                "block_id": "design_block",
                "label": {"type": "plain_text", "text": "Design"},
                "element": {
                    "type": "plain_text_input",
                    "action_id": "val",
                    "initial_value": lane.get("design") or "",
                    "placeholder": {"type": "plain_text",
                                    "text": "e.g. Tivoli" if which == "field" else "e.g. Greek Key"},
                },
            },
            {
                "type": "input",
                "block_id": "difficulty_block",
                "label": {"type": "plain_text", "text": "Difficulty"},
                "element": {
                    "type": "plain_text_input",
                    "action_id": "val",
                    "max_length": 2,
                    "initial_value": lane.get("difficulty") or "",
                    "placeholder": {"type": "plain_text", "text": DIFFICULTY_HINT},
                },
            },
        ],
    }


@app.view("trk_lane_details")
def handle_lane_details(ack, body, client):
    """Save what the lane is, then start the work the assembler pressed for."""
    vals = body["view"]["state"]["values"]
    meta = json.loads(body["view"]["private_metadata"])

    design = (_typed(vals, "design_block", "val") or "").strip()
    if not design:
        ack(response_action="errors", errors={"design_block": "Name the design."})
        return
    difficulty, difficulty_error = read_difficulty(_typed(vals, "difficulty_block", "val"))
    if difficulty_error:
        ack(response_action="errors", errors={"difficulty_block": difficulty_error})
        return
    if not difficulty:
        ack(response_action="errors", errors={"difficulty_block": "Give it a difficulty."})
        return

    ack()
    user_id = body["user"]["id"]
    task_id = meta["task_id"]
    channel_id = meta["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return
    # The form may have sat open while the assembler resumed another job from its
    # card. Nothing is recorded for a press that cannot start.
    if busy_elsewhere(client, task, user_id, channel_id):
        return

    # The details are written against the lane whichever it is - the same call
    # the border form has always made, now told which part and which lane. No
    # jig: it is recorded from the card, once the assembler knows it.
    database.set_lane_details(
        task_id, meta["phase"], design, difficulty, part=meta["part"],
    )

    outcome = database.start_work(
        task_id, meta["phase"], meta["activity"] or "production", part=meta["part"],
    )
    if outcome != "started":
        client.chat_postEphemeral(
            channel=channel_id, user=user_id,
            text=refusal_text(outcome, task, meta["phase"]),
        )
        update_card(client, database.get_task(task_id), channel_id)
        return

    update_card(client, database.get_task(task_id), channel_id)


@app.action("trk_stop_task")
def handle_stop(ack, body, client):
    """
    Pause. The assembler is not working on this job for the moment.

    Whatever was being timed stops, including any cutting that was being
    measured inside it - there is nothing left for cutting to be inside. The
    job keeps everything it has recorded and nothing about it is finished.
    """
    ack()
    task_id = read_work_value(body["actions"][0]["value"])[0]
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return

    database.stop_work(task_id)
    update_card(client, database.get_task(task_id), channel_id)


@app.action("trk_start_cutting")
def handle_start_cutting(ack, body, client):
    """
    The assembler goes and cuts tiles for a while.

    The sheeting timer keeps running, because they are still working this job -
    they have gone downstairs to cut for it. This measures how much of that
    time was spent cutting; it never takes time away from the sheeting.

    It is never told which part. Cutting is measured INSIDE the running
    segment, and that segment already names the part - so the attribution
    comes from the work it is happening in rather than from the button.
    """
    ack()
    task_id = read_work_value(body["actions"][0]["value"])[0]
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return

    outcome = database.start_cutting(task_id)
    if outcome != "started":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=refusal_text(outcome, task, task["current_phase"]),
        )
        return

    update_card(client, database.get_task(task_id), channel_id)


@app.action("trk_stop_cutting")
def handle_stop_cutting(ack, body, client):
    """Back upstairs. The sheeting was running throughout and still is."""
    ack()
    task_id = read_work_value(body["actions"][0]["value"])[0]
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return

    outcome = database.stop_cutting(task_id)
    if outcome != "stopped":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=refusal_text(outcome, task, task["current_phase"]),
        )
        return

    update_card(client, database.get_task(task_id), channel_id)


@app.action("trk_start_packing")
def handle_start_packing(ack, body, client):
    """
    Go and pack for a while, leaving the sheeting where it is.

    New cards put the other work on the card itself. This stays registered
    because a card posted by an earlier version of the tracker is still live in
    somebody's DM, and it should keep working rather than fall silent the
    moment a deployment lands.
    """
    ack()
    task_id = read_work_value(body["actions"][0]["value"])[0]
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return

    outcome = database.start_packing(task_id)
    if outcome != "started":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=refusal_text(outcome, task, "packing"),
        )
        return

    update_card(client, database.get_task(task_id), channel_id)


# ---------------------------------------------------------------------------
# Finishing
# ---------------------------------------------------------------------------
# Two different presses share one action id, told apart by what the button
# carries. A lane's Finish names that lane and closes it, and nothing else
# happens - no form, no channel post, no move to somewhere the assembler did not
# ask to go. The job's Finish carries only the job and appears once nothing
# anywhere on it is unfinished; that one opens the closing notes.

@app.action("trk_complete_task")
def handle_complete(ack, body, client):
    """Finish one lane, or - when the job is done - open the closing notes."""
    ack()
    task_id, part, phase, _activity = read_work_value(body["actions"][0]["value"])
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return

    if phase is None:
        # The job itself. Offered only when every lane on every part is done,
        # and checked again here because the card may have been sitting open.
        if not job_is_finishable(task):
            client.chat_postEphemeral(
                channel=channel_id, user=user_id,
                text=("T-" + str(task_id) + " still has work on it. Finish each part's work "
                      "first, then the job."),
            )
            return
        client.views_open(
            trigger_id=body["trigger_id"],
            view=notes_modal_view(
                json.dumps({
                    "task_id": task_id,
                    "dm_channel_id": channel_id,
                    "team_channel_id": task["channel_id"],
                }),
                "*Everything on this job is finished.* Anything worth recording before it "
                "closes?",
            ),
        )
        return

    outcome = database.complete_task(task_id, phase=phase, part=part)
    if outcome != "completed":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=refusal_text(outcome, task, phase),
        )
        return

    # The card, and nothing else. Finishing a lane is the assembler's business and
    # the channel does not hear about it; what comes next is on the card, which
    # now offers whatever is still unfinished.
    update_card(client, database.get_task(task_id), channel_id)


@app.view("trk_notes_modal")
def handle_notes_submission(ack, body, client):
    ack()
    user_id = body["user"]["id"]
    vals = body["view"]["state"]["values"]
    metadata = json.loads(body["view"]["private_metadata"])
    task_id = metadata["task_id"]
    dm_channel_id = metadata["dm_channel_id"]
    team_channel_id = metadata["team_channel_id"]

    general_notes = vals["notes_block"]["general_notes"]["value"] or "None"
    issues = vals["issues_block"]["issues"]["value"] or "None"

    database.save_notes_and_complete(task_id, general_notes, issues)
    task = database.get_task(task_id)

    total_time = database.format_duration(task["total_elapsed"])

    # The assembler's own card keeps the full record: it is their work, in their
    # DM, and the breakdown is the thing they would want to look back at.
    client.chat_update(
        channel=dm_channel_id,
        ts=task["message_ts"],
        text=f"T-{task_id} is finished.",
        blocks=(
            [
                {
                    "type": "header",
                    "text": {"type": "plain_text", "text": header_text(task, "  -  finished")},
                },
                {
                    "type": "section",
                    "text": {"type": "mrkdwn", "text": f"*Total time: {total_time}*"},
                },
            ]
            + [
                {"type": "section", "text": {"type": "mrkdwn", "text": "\n".join(_time_lines(task))}}
            ]
        ),
    )

    # THE PUBLIC CHANNEL CARRIES NO TIMING. It says what happened and who did
    # it, and nothing about how long anything took. A workshop channel is an
    # operational feed, not a performance board, and a breakdown of one
    # person's minutes posted where everyone reads it is a different thing from
    # the same breakdown on their own card.
    client.chat_postMessage(
        channel=team_channel_id,
        text=f"{MARK_FINISHED} <@{user_id}> has finished T-{task_id} {task['customer_name']}",
    )


# The jig. A lane sometimes needs another one part way through: maybe the first
# turned out wrong and was swapped, maybe two sizes are genuinely needed
# together. Either way the earlier jig really was used, so this ADDS a record
# next to it - it never overwrites one. Typing mistakes are fixed through Edit
# instead, which changes the value it names.
#
# It lives on every working card because the jig is normally established during
# the setup, which is after the job was logged and can be well after the
# sheeting started.

# ---------------------------------------------------------------------------
# Jig and template
# ---------------------------------------------------------------------------
# A phase records any number of jigs, in the order they were used. Adding
# appends, because a jig that was genuinely used stays on the record; a
# typing mistake is corrected through Edit instead.

@app.action("trk_add_jig")
def handle_add_jig(ack, body, client):
    """
    Record the jig on the work the assembler is doing.

    It asks nothing but the value. The button is only on the card while a lane
    is being worked, and it carries which part and which lane - so there is no
    "which work used it?" left to ask. That question existed because the jig
    button used to sit on every card, including ones where the answer was not
    obvious; now the card only offers it where the answer is the work in hand.
    """
    ack()
    task_id, part, phase, _activity = read_work_value(body["actions"][0]["value"])
    user_id = body["user"]["id"]
    channel_id = body["container"]["channel_id"]

    task = resolve_job(client, body, task_id, user_id, channel_id)
    if task is None:
        return

    here = task.get("working_on") or {}
    phase = phase or here.get("phase")
    if part is None:
        part = here.get("part")
    if phase not in ("field_sheeting", "border_sheeting"):
        client.chat_postEphemeral(
            channel=channel_id, user=user_id,
            text="Start the field or the border first - a jig belongs to the work it is used on.",
        )
        return

    named = work_name(phase, "production", part_label(task, part))
    client.views_open(
        trigger_id=body["trigger_id"],
        view={
            "type": "modal",
            "callback_id": "trk_add_jig_modal",
            "title": {"type": "plain_text", "text": "Jig or template"},
            "submit": {"type": "plain_text", "text": "Save"},
            "close": {"type": "plain_text", "text": "Cancel"},
            "private_metadata": json.dumps({
                "task_id": task_id,
                "channel_id": channel_id,
                "part": part,
                "phase": phase,
            }),
            "blocks": [
                {
                    "type": "input",
                    "block_id": "jig_block",
                    "label": {"type": "plain_text", "text": "Jig or template for " + lower_name(named)},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "jig_size",
                        "placeholder": {"type": "plain_text",
                                        "text": "e.g. 49.6, 49.4/49.8, or template"},
                    },
                },
            ],
        },
    )


@app.view("trk_add_jig_modal")
def handle_add_jig_submission(ack, body, client):
    ack()
    user_id = body["user"]["id"]
    vals = body["view"]["state"]["values"]
    metadata = json.loads(body["view"]["private_metadata"])
    task_id = metadata["task_id"]
    channel_id = metadata["channel_id"]

    jig_size = (vals["jig_block"]["jig_size"]["value"] or "").strip()
    if not jig_size:
        return

    database.add_jig(task_id, metadata.get("phase") or "field_sheeting", jig_size,
                     part=metadata.get("part"))
    task = database.get_task(task_id)

    # The job can disappear between opening the modal and submitting it -
    # deleted, or finished. Tell the assembler instead of closing the modal in
    # silence with their jig unrecorded, and leave the final card alone.
    if task is None or task["current_phase"] == "completed":
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text=f"'{jig_size}' was not recorded - that job is no longer open."
        )
        return

    update_card(client, task, channel_id, note=f"*Jig recorded: {jig_size}*")


#Delete Button
# ---------------------------------------------------------------------------
# Corrections: editing a job, and cancelling one
# ---------------------------------------------------------------------------
# Editing changes the values a job was given. Cancelling keeps the job and
# everything recorded on it - nothing is ever deleted outright.

@app.action("trk_delete_task")
def handle_delete(ack, body, client):
    ack()
    task_id, part, _phase, _activity = read_work_value(body["actions"][0]["value"])
    user_id = body["user"]["id"]
    task = database.get_task(task_id)
    channel_id = body["container"]["channel_id"]

    if task is None:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="That job is no longer on your list."
        )
        return

    # A job is only ever taken off the list by the assembler whose job it is. There
    # is no supervisor path anywhere in the tracker yet, so this check is the
    # whole of the rule - it is not a friendly message in front of a second one
    # enforced further down.
    if task["user_id"] != user_id:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="You can only take your own jobs off the list."
        )
        return

    # Delete from the database
    database.delete_task(task_id)

    # The card is replaced by what actually happened. Nothing is destroyed -
    # LMSA cancels the job and keeps it, which is exactly why an assembler may safely
    # use this on a job that turned out to be a mistake. Saying "deleted" here
    # contradicted the dialog they had just agreed to, one press earlier.
    client.chat_update(
        channel=channel_id,
        ts=task["message_ts"],
        text=f"Job T-{task_id} was cancelled by <@{user_id}>. Its recorded time was kept.",
        blocks=[
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": (f"*Job T-{task_id} was cancelled by <@{user_id}>.*"
                             "\nIts recorded time and history have been kept.")
                }
            }
        ]
    )


@app.action("trk_edit_task")
def handle_edit(ack, body, client):
    ack()
    task_id, part, _phase, _activity = read_work_value(body["actions"][0]["value"])
    user_id = body["user"]["id"]
    task = database.get_task(task_id)
    channel_id = body["container"]["channel_id"]

    if task is None:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="That job is no longer on your list."
        )
        return

    if task["user_id"] != user_id:
        client.chat_postEphemeral(
            channel=channel_id,
            user=user_id,
            text="You can only edit your own jobs."
        )
        return

    # Bundle task_id and channel_id to pass through the modal
    edit_metadata = json.dumps({
        "task_id": task_id,
        "channel_id": channel_id
    })

    # One box per jig already recorded, pre-filled, so a mistyped value can
    # be corrected later - even after that phase has finished. These boxes
    # fix EXISTING jigs; a genuinely new jig goes through Add Jig instead.
    jig_blocks = []
    for i, rec in enumerate(task["field_jig_records"], start=1):
        jig_blocks.append({
            "type": "input",
            "block_id": f"jig_edit_{rec['id']}",
            "label": {"type": "plain_text", "text": f"Field jig / template {i}"},
            "element": {
                "type": "plain_text_input",
                "action_id": "jig_value",
                "initial_value": rec["value"]
            }
        })
    for i, rec in enumerate(task["border_jig_records"], start=1):
        jig_blocks.append({
            "type": "input",
            "block_id": f"jig_edit_{rec['id']}",
            "label": {"type": "plain_text", "text": f"Border jig / template {i}"},
            "element": {
                "type": "plain_text_input",
                "action_id": "jig_value",
                "initial_value": rec["value"]
            }
        })

    # Open pre-filled edit modal
    client.views_open(
        trigger_id=body["trigger_id"],
        view={
            "type": "modal",
            "callback_id": "trk_edit_task_modal",
            "title": {"type": "plain_text", "text": "Edit job"},
            "submit": {"type": "plain_text", "text": "Save changes"},
            "close": {"type": "plain_text", "text": "Cancel"},
            "private_metadata": edit_metadata,
            # The same words as the intake form and the card: sentence case,
            # and the names the workshop already uses for these things.
            "blocks": [
                {
                    "type": "input",
                    "block_id": "customer_block",
                    "label": {"type": "plain_text", "text": "Customer name"},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "customer_name",
                        "initial_value": task["customer_name"]
                    }
                },
                {
                    "type": "input",
                    "block_id": "invoice_block",
                    "label": {"type": "plain_text", "text": "Invoice / Pro Forma number"},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "invoice_num",
                        "initial_value": task["invoice_number"]
                    }
                },
                {
                    "type": "input",
                    "block_id": "task_block",
                    "label": {"type": "plain_text", "text": "Job description"},
                    "element": {
                        "type": "plain_text_input",
                        "multiline": True,
                        "action_id": "task_desc",
                        "initial_value": task["task_description"]
                    }
                },
                {
                    "type": "input",
                    "block_id": "design_block",
                    "label": {"type": "plain_text", "text": "Field design"},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "design",
                        "initial_value": task["field_design"] or ""
                    }
                },
                {
                    "type": "input",
                    "block_id": "difficulty_block",
                    "label": {"type": "plain_text", "text": DIFFICULTY_LABEL_FIELD},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "difficulty",
                        "max_length": 2,
                        "placeholder": {"type": "plain_text", "text": DIFFICULTY_HINT},
                        "initial_value": task["difficulty"] or ""
                    }
                },
                *jig_blocks,
                {
                    "type": "input",
                    "block_id": "date_block",
                    "optional": True,
                    # The same field as the intake form, so the same words, the
                    # same format and the same rules. Both build the box from
                    # DUE_DATE_LABEL and DUE_DATE_HINT so one box cannot end up
                    # promising two different things.
                    "label": {"type": "plain_text", "text": DUE_DATE_LABEL},
                    "element": {
                        "type": "plain_text_input",
                        "action_id": "due_date",
                        "placeholder": {"type": "plain_text", "text": DUE_DATE_HINT},
                        "initial_value": due_date_supplied(task) or ""
                    }
                }
            ]
        }
    )


@app.view("trk_edit_task_modal")
def handle_edit_submission(ack, body, client):
    user_id = body["user"]["id"]
    vals = body["view"]["state"]["values"]

    # Retrieve task_id and channel_id from metadata
    metadata = json.loads(body["view"]["private_metadata"])
    task_id = metadata["task_id"]
    channel_id = metadata["channel_id"]

    # Collect updated values
    customer_name = vals["customer_block"]["customer_name"]["value"]
    invoice_number = vals["invoice_block"]["invoice_num"]["value"]
    task_description = vals["task_block"]["task_desc"]["value"]
    design = vals["design_block"]["design"]["value"]
    difficulty, difficulty_error = read_difficulty(
        vals["difficulty_block"]["difficulty"]["value"])
    if difficulty_error:
        ack(response_action="errors", errors={"difficulty_block": difficulty_error})
        return
    # What the job said before the modal opened. Read first, because the due
    # date needs it too.
    task_before = database.get_task(task_id)

    # Same box, same rules as the intake form. One exception, and it is about
    # history rather than about dates: a row written before this screen asked
    # for a real date may hold free text, which the form pre-fills. Judging
    # that on submit would stop an assembler fixing a customer's name until they had
    # also rewritten a due date they never touched. So a value that comes back
    # exactly as it was stored passes through untouched - that is not new
    # typing, and nothing about it is being changed.
    typed_due_date = vals["date_block"]["due_date"]["value"]
    stored_due_date = (task_before["due_date"] if task_before else None) or ""
    if stored_due_date.strip() and (typed_due_date or "").strip() == stored_due_date.strip():
        due_date = stored_due_date
    else:
        due_date, due_date_error = read_due_date(typed_due_date)
        if due_date_error:
            ack(response_action="errors", errors={"date_block": due_date_error})
            return
    ack()

    # What each jig said before the modal opened, so only boxes the assembler
    # actually changed get corrected
    previous_jigs = {}
    if task_before is not None:
        for rec in task_before["field_jig_records"] + task_before["border_jig_records"]:
            previous_jigs[rec["id"]] = rec["value"]

    # Save to database
    database.update_task(
        task_id=task_id,
        customer=customer_name,
        invoice=invoice_number,
        task_desc=task_description,
        design=design,
        difficulty=difficulty,
        due_date=due_date
    )

    # Fix any jig boxes the assembler changed. Each correction names its own
    # record, so fixing one jig never touches the others.
    for block_id, entry in vals.items():
        if not block_id.startswith("jig_edit_"):
            continue
        jig_id = block_id[len("jig_edit_"):]
        new_value = (entry["jig_value"]["value"] or "").strip()
        if new_value and new_value != previous_jigs.get(jig_id):
            database.correct_jig(task_id, jig_id, new_value)

    # Refresh the card so the corrections show. The job's own state decides
    # what it says and which buttons it keeps - an edit is not a change of
    # what the assembler is doing.
    task = database.get_task(task_id)
    if task is not None and task["current_phase"] != "completed":
        update_card(client, task, channel_id, note="*Details updated.*")


# ---------------------------------------------------------------------------
# Running the tracker on its own
# ---------------------------------------------------------------------------
# Inside LMSA the relay delivers events instead, so nothing below runs there.
# This is the path for running the tracker standalone against Slack.

if __name__ == "__main__":
    database.setup_database()
    handler = SocketModeHandler(app, os.environ.get("SLACK_APP_TOKEN"))
    print("Trackbot is running!")
    handler.start()
